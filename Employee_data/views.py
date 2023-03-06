import openpyxl
from django.shortcuts import render
from .models import EmployeeData,Expenses
from django.views.generic import View
from django.http import request,HttpResponse
from openpyxl import Workbook,load_workbook,writer,drawing
import datetime
import pandas as pd
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
import win32com.client
from pywintypes import *
import os
import pdfkit
# Create your views here.
def home(request):
    return render(request, "welcome.html")
def employeeDetails(request):
    empdetails = EmployeeData.objects.all()
    return render(request, 'Employee_details.html', {"empdetails":empdetails})

class getEmployeeById(View):
    def get(self,request):
        return render(request,'getemployeedetails.html')
    def post(self,request):
        try:
            id = request.POST['eid']
            employeeRecord = EmployeeData.objects.get(Employee_ID=id)
            print(employeeRecord.Employee_Date_of_joined)
            Dateofjoined = employeeRecord.Employee_Date_of_joined.strftime('%m-%d-%Y')
            date = datetime.datetime.now() - datetime.timedelta(days=20)
            month = date.month
            month_number = "%d" % month
            l=[['1','3','5','7','8','10','12'],['2'],['4','6','9','11']]
            if month_number in l[0]:
                actualdays = 31
            elif month_number in l[1]:
                actualdays = 28
            elif month_number in l[2]:
                actualdays = 30
            return render(request,'empsal.html',{'employeeRecord':employeeRecord,'dateofjoined':Dateofjoined,'actualdays':actualdays})
        except EmployeeData.DoesNotExist:
            return render(request,'getemployeedetails.html',{'msg':'does not exist'})

def Salary(request):

    id=request.POST["EmpID"]
    firstname=request.POST["EmpFirstname"]
    middlename = request.POST["EmpMiddlename"]
    lastname = request.POST["EmpLastname"]
    department = request.POST["EmpDepartment"]
    designation = request.POST["EmpDesignation"]
    dateofjoined = request.POST["EmpDateofjoined"]
    bankname= request.POST["EmpBankname"]
    accountnumber = request.POST["EmpBankaccountnumber"]
    ifsccode = request.POST["EmpIfsccode"]
    pannumber = request.POST["EmpPannumber"]
    uannumber = request.POST["EmpUannumber"]
    pfnumber = request.POST["EmpPfnumber"]
    esinumber =request.POST["EmpEsinumber"]
    actualworkingdays = request.POST["EmpActualworkingdays"]
    totalworkingdays = request.POST["EmpTotalworkingdays"]
    lossofdays = request.POST["EmpLossofpaydays"]
    paidleaves = request.POST["EmpPaidleaves"]
    dayspayable= request.POST["EmpDayspayable"]
    basic = request.POST["EmpBasic"]
    coveyanceallowance = request.POST["EmpConveyanceallowance"]
    hra = request.POST["EmpHra"]
    medicalallowance = request.POST["EmpMedicalallowance"]
    specialallowance=request.POST["EmpSpecialallowance"]
    variablepay = request.POST["EmpVariablepay"]
    totalearnings = request.POST["EmpTotalearnings"]
    pfemployee= request.POST["EmpPfemployee"]
    pfemployer = request.POST["EmpPfemployer"]
    esiemployee =request.POST["EmpEsiemployee"]
    esiemployer = request.POST["EmpEsiemployer"]
    totalcontributions=request.POST["EmpTotalcontributions"]
    professionaltax=request.POST["EmpProfessionaltax"]
    totaldeductions =request.POST["EmpTotaldeductions"]
    netpayable = request.POST["EmpNetpayable"]
    path = "C:\\Users\\hp\\Desktop\\new\\payment\\Payment_system\\Employee_data\\static\\excel\\myfile.xlsx"
    ref_workbook = openpyxl.load_workbook(path)
    date = datetime.datetime.now() - datetime.timedelta(days=20)
    month = date.month
    month_number = "%d" % month
    datetime_object = datetime.datetime.strptime(month_number, "%m")
    month_name = datetime_object.strftime("%b")
    print(month_name)
    year = date.year
    wb=ref_workbook.get_sheet_names()
    sheet=ref_workbook.get_sheet_by_name('Sheet1')
    sheet["A1"] = "PAYSLIP %s %d" % (month_name,year)
    sheet["A8"] = "%s %s %s" %(firstname,middlename,lastname)
    sheet["A11"] =id
    sheet["C11"]=dateofjoined
    sheet["E11"]=department
    sheet["G11"]="-"
    sheet["A14"]=designation
    sheet["C14"]="Bank Transfer"
    sheet["E14"]=bankname
    sheet["G14"]=ifsccode
    sheet["A17"]=accountnumber
    sheet["C17"]=pannumber
    sheet["E17"]=uannumber
    sheet["G17"]=pfnumber
    sheet["A20"]=esinumber
    sheet["A26"]=actualworkingdays
    sheet["C26"]=totalworkingdays
    sheet['E26']=lossofdays
    sheet["G26"]=dayspayable
    sheet["C29"]=basic
    sheet["C30"]=coveyanceallowance
    sheet["c31"]=hra
    sheet["c32"]=medicalallowance
    sheet["C33"]=specialallowance
    sheet["C34"]=variablepay
    sheet["C35"]=totalearnings
    sheet["G29"]=pfemployee
    sheet["G30"]=pfemployer
    sheet["G31"]=esiemployee
    sheet["G32"]=esiemployer
    sheet["G33"]=totalcontributions
    sheet["G37"]=professionaltax
    sheet["G38"]=totaldeductions
    sheet["C40"]=netpayable
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         # check if the cell value is a string
    #         if isinstance(cell.value, str):
    #             # try to convert the string to a float
    #             try:
    #                 cell.value = float(cell.value)
    #             except ValueError:
    #                 continue



    # pdf_filename = ("%s %s-Payslip-%s-%d.xlsx" %(firstname,lastname,month_name,year))
    # pdf = canvas.Canvas(pdf_filename)
    #
    # # Iterate through the rows and columns of the XLSX file
    # for row in sheet.rows:
    #     for cell in row:
    #         # Get the cell value and write it to the PDF file
    #         pdf.drawString(cell.column_letter + str(cell.row), str(cell.value), str(pdf_filename) )
    #
    # # Save the PDF file
    # pdf.save()
    #
    # print('XLSX file has been converted to PDF successfully!')
    # list=[]
    file = ("%s %s-Payslip-%s-%d.xlsx" %(firstname,lastname,month_name,year))
    # print(file)
    # sheet.save(file)
    # ref_workbook.save("%s %s-Payslip-%s-%d.xlsx" %(firstname,lastname,month_name,year))
    # cessfully completed")
    # ref_workbook.ExportasFixedFormat("%s %s-Payslip-%s-%d.pdf" %(firstname,lastname,month_name,year))
    ref_workbook.save(f'C:\\Users\\hp\\Desktop\\new\\payment\\Payment_system\\payslip\\{file}')
    print(date)

    return render(request,"success.html")


#

class ExpensesCompany(View):
    def get(self,request):
        return render(request,'addexpenses.html')
    def post(self,request):
        bill_number = request.POST["billnumber"]
        bill_price = request.POST["billprice"]
        bill_date = request.POST["billdate"]
        billimage = request.Files["bill_image"]
        Expenses_record = Expenses(billNo=bill_number, price=bill_price, Date=bill_date, bill_image=billimage)
        Expenses_record.save()
        return render(request,'status.html')


def get_files_in_folder(folder_path):
    files = []
    for file_name in os.listdir(folder_path):
        if os.path.isfile(os.path.join(folder_path, file_name)):
            files.append(file_name)
    return files

# def my_view(request):
#     folder_path = 'F:\payment\Payment_system\payslip'
#     files = get_files_in_folder(folder_path)
#     print(files)
#     for file in files:
        # import win32com.client
        # from pywintypes import com_error
        # WB_PATH = file
        # PATH_TO_PDF = f'F:\\payment\\Payment_system\\pdf salary\\{file}.pdf'
        # excel = win32com.client.Dispatch('Excel.Application')
        # excel.Visible = False
        # try:
        #     print('Start conversion to PDF')
        #     Workbooks = excel.Workbooks.Open(WB_PATH)
        #     Workbooks.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
        # finally:
        #     print('Succeeded.')
        # o = win32com.client.Dispatch("Excel.Application")
        # o.Visible = False
        # o.DisplayAlerts = False
        # wb = o.Workbooks.Open(file)
        # wb.WorkSheets("sheet1").Select()
        # wb.ActiveSheet.ExportAsFixedFormat(0, f"{file}.pdf")
        # o.Quit()
        # from PDFWriter import PDFWriter
    #     from PyPDF2 import PdfFileReader, PdfFileWriter,PdfWriter
    #
    #     workbook = load_workbook(file, data_only=True)
    #     worksheet = workbook.active
    #
    #     pw = PdfWriter('fruits2.pdf')
    #     # pw.setfont('Courier', 12)
    #     # pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
    #     # pw.setFooter('Generated using openpyxl and xtopdf')
    #
    #     #
    #
    # return render(request,'pdffiles.html')


import io
from django.http import FileResponse
from django.template.loader import get_template
from django.views import View
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import openpyxl
from django.http import HttpResponse

import win32com.client
import pythoncom
class my_view(View):
    def get(self, request):
        folder_path = 'C:\\Users\\hp\\Desktop\\new\\payment\\Payment_system\\payslip'
        files = get_files_in_folder(folder_path)
        print(files)
        WB_PATH = f'C:\\Users\\hp\\Desktop\\new\\payment\\Payment_system\\payslip\\{files[2]}'
        PATH_TO_PDF = f'C:\\Users\\hp\\Desktop\\new\\payment\\Payment_system\\pdf salary\\{files[2]}.pdf'
        excel = win32com.client.Dispatch("Excel.Application",pythoncom.CoInitialize())
        excel.Visible = False
        try:
            print('Start conversion to PDF')
            Workbooks = excel.Workbooks.Open(WB_PATH)
            Workbooks.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
            Workbooks.Close(False)
            excel.Quit()


        finally:
            print('Succeeded.')
        return render(request, 'status.html')
            # Convert the worksheet to HT
        # pdf_filename = (f'{file}.pdf')
        # pdf = canvas.Canvas(pdf_filename)
        #
        # # Iterate through the rows and columns of the XLSX file
        # for row in worksheet.rows:
        #     for cell in row:
        #         # Get the cell value and write it to the PDF file
        #         pdf.drawString(cell.column_letter + str(cell.row), str(cell.value), str(pdf_filename) )
        #
        # # Save the PDF file
        # pdf.save()